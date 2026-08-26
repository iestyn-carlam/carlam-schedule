#!/usr/bin/env python3
"""
Pulls every row from the Carlam Team Schedule Notion database and writes a
single pivoted Excel workbook (docs/schedule-grid.xlsx) that looks like the
old Excel layout: one column per person, one row per date, each cell showing
that person's status/task for that day.

This runs alongside generate_ics.py in the same GitHub Actions workflow, so
every sync updates both the individual Outlook calendar feeds AND this one
shared grid file, from the same Notion data.

No paid service is involved anywhere in this pipeline.
"""

import hashlib
import os
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Reuses the same private salt as generate_ics.py, so this file's name is
# just as unguessable as the individual calendar links. Without it set, the
# script refuses to run rather than silently using a public/predictable name.
FILENAME_SALT = os.environ.get("FILENAME_SALT")
if not FILENAME_SALT:
    print(
        "FILENAME_SALT is not set. Refusing to run: without a private salt, "
        "the grid filename would be guessable. Add a FILENAME_SALT repository "
        "secret in GitHub (Settings -> Secrets and variables -> Actions) and "
        "re-run.",
        file=sys.stderr,
    )
    sys.exit(1)

NOTION_TOKEN = os.environ["NOTION_TOKEN"]
DATABASE_ID = os.environ["NOTION_DATABASE_ID"]
NOTION_VERSION = "2022-06-28"
OUTPUT_DIR = Path(__file__).parent / "docs"
GRID_SUFFIX = hashlib.sha256(f"{FILENAME_SALT}:schedule-grid".encode()).hexdigest()[:8]
OUTPUT_FILE = OUTPUT_DIR / f"schedule-grid-{GRID_SUFFIX}.xlsx"

HEADERS = {
    "Authorization": f"Bearer {NOTION_TOKEN}",
    "Notion-Version": NOTION_VERSION,
    "Content-Type": "application/json",
}

# Colour-code the most common statuses so the grid is scannable at a glance,
# similar to how the old spreadsheet used cell shading.
STATUS_COLOURS = {
    "A/L": "FFF2CC",
    "OFF": "F4CCCC",
    "TOIL": "F4CCCC",
    "Offline Edit": "D9EAD3",
    "Online Edit": "D9EAD3",
    "Finishing": "CFE2F3",
    "Delivery": "C9DAF8",
    "Archive": "EFEFEF",
    "Priority task": "F9CB9C",
    "Notes / Changes": "FFF2CC",
    "Tentative": "EAD1DC",
}
DEFAULT_COLOUR = "FFFFFF"


def fetch_all_rows():
    """Query the Notion database, following pagination until done."""
    rows = []
    payload = {"page_size": 100}
    url = f"https://api.notion.com/v1/databases/{DATABASE_ID}/query"

    while True:
        resp = requests.post(url, headers=HEADERS, json=payload)
        resp.raise_for_status()
        data = resp.json()
        rows.extend(data["results"])
        if not data.get("has_more"):
            break
        payload["start_cursor"] = data["next_cursor"]

    return rows


def get_plain_text(rich_text_list):
    return "".join(t.get("plain_text", "") for t in rich_text_list) if rich_text_list else ""


def extract_row(page):
    props = page["properties"]

    title = get_plain_text(props.get("Task", {}).get("title", []))
    date_prop = props.get("Date", {}).get("date")
    start_date = date_prop["start"][:10] if date_prop else None

    programme = (props.get("Programme", {}).get("select") or {}).get("name", "")
    status = (props.get("Status", {}).get("status") or props.get("Status", {}).get("select") or {}).get("name", "")
    notes = get_plain_text(props.get("Notes", {}).get("rich_text", []))

    people = [p.get("name", "Unknown") for p in props.get("Person", {}).get("people", [])]
    person_name_prop = props.get("Person Name", {})
    if "select" in person_name_prop:
        person_name_text = (person_name_prop.get("select") or {}).get("name", "")
    else:
        person_name_text = get_plain_text(person_name_prop.get("rich_text", []))
    if not people and person_name_text:
        people = [person_name_text]

    return {
        "title": title,
        "start_date": start_date,
        "programme": programme,
        "status": status,
        "notes": notes,
        "people": people,
    }


def cell_text(row) -> str:
    """What goes in a single grid cell for one person on one day."""
    parts = [p for p in [row["status"], row["programme"] or row["title"]] if p]
    return " - ".join(parts) if parts else (row["title"] or "")


def build_workbook(rows):
    # Collect every person and every date that actually appears.
    people = sorted({p for r in rows for p in r["people"]})
    dated_rows = [r for r in rows if r["start_date"]]

    if not dated_rows:
        min_date = max_date = date.today()
    else:
        all_dates = [date.fromisoformat(r["start_date"]) for r in dated_rows]
        min_date = min(all_dates)
        max_date = max(all_dates)

    # Grid: (date, person) -> list of cell strings/colours (usually one entry,
    # but a person could in theory have two things tagged the same day).
    grid = defaultdict(list)
    for r in dated_rows:
        d = date.fromisoformat(r["start_date"])
        for person in r["people"]:
            grid[(d, person)].append(r)

    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4A4A4A")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    weekend_fill = PatternFill("solid", fgColor="F5F5F5")

    # Header row: blank corner, then one column per person.
    ws.cell(row=1, column=1, value="Date").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).border = border
    for col, person in enumerate(people, start=2):
        c = ws.cell(row=1, column=col, value=person)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = Alignment(horizontal="center")

    # One row per calendar day in range, even if nothing's scheduled, so the
    # grid reads continuously like the old spreadsheet did.
    current = min_date
    excel_row = 2
    while current <= max_date:
        is_weekend = current.weekday() >= 5
        date_cell = ws.cell(row=excel_row, column=1, value=current.strftime("%a %d %b %Y"))
        date_cell.font = Font(bold=True)
        date_cell.border = border
        if is_weekend:
            date_cell.fill = weekend_fill

        for col, person in enumerate(people, start=2):
            entries = grid.get((current, person), [])
            cell = ws.cell(row=excel_row, column=col)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if is_weekend and not entries:
                cell.fill = weekend_fill
            if entries:
                text = "\n".join(cell_text(e) for e in entries)
                cell.value = text
                colour = STATUS_COLOURS.get(entries[0]["status"], DEFAULT_COLOUR)
                cell.fill = PatternFill("solid", fgColor=colour)

        current += timedelta(days=1)
        excel_row += 1

    # Column widths / freeze panes so it behaves like a normal spreadsheet.
    ws.column_dimensions["A"].width = 20
    for col in range(2, len(people) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 22
    ws.freeze_panes = "B2"

    return wb


def main():
    print("Fetching rows from Notion...")
    pages = fetch_all_rows()
    rows = [extract_row(p) for p in pages]
    print(f"Fetched {len(rows)} rows.")

    wb = build_workbook(rows)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_FILE)
    print(f"Wrote {OUTPUT_FILE}")


if __name__ == "__main__":
    try:
        main()
    except requests.HTTPError as e:
        print(f"Notion API error: {e.response.status_code} {e.response.text}", file=sys.stderr)
        sys.exit(1)
